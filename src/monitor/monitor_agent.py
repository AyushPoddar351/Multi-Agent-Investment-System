"""
Monitor Agent - Pure Data Collection & Analysis
Multi-Agent Investment Management System - Refactored Version
"""

import yfinance as yf
import pandas as pd
import numpy as np
import json
from datetime import datetime, timedelta
from typing import Dict, List, Tuple, Optional
from dataclasses import dataclass
import os
import sys
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import LineChart, BarChart, PieChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
import warnings
warnings.filterwarnings('ignore')

# Add parent directory to path for imports
sys.path.append(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
from logger.custom_logger import CustomLogger

# Configure custom logger
logger = CustomLogger().get_logger(__file__)

@dataclass
class StockData:
    """Pure stock data without recommendations"""
    symbol: str
    company_name: str
    sector: str
    current_price: float
    price_change: float
    price_change_pct: float
    volume: int
    market_cap: float
    pe_ratio: float
    open_price: float
    high_price: float
    low_price: float
    date: str

@dataclass
class TechnicalIndicators:
    """Technical analysis indicators"""
    symbol: str
    rsi: float
    sma_20: float
    sma_50: float
    sma_200: float
    volatility: float
    momentum_20d: float
    momentum_50d: float
    bollinger_upper: float
    bollinger_lower: float
    macd: float
    signal_strength: str  # STRONG_UP, UP, NEUTRAL, DOWN, STRONG_DOWN

@dataclass
class SectorAnalysis:
    """Sector performance analysis"""
    sector: str
    stock_count: int
    avg_price_change: float
    avg_volume_change: float
    sector_volatility: float
    trend_direction: str  # UPTREND, DOWNTREND, SIDEWAYS
    top_performers: List[str]
    bottom_performers: List[str]
    sector_strength: str  # STRONG, MODERATE, WEAK

@dataclass
class MarketOverview:
    """Overall market condition assessment"""
    analysis_date: datetime
    market_sentiment: str  # BULLISH, BEARISH, NEUTRAL
    market_volatility: float
    advancing_stocks: int
    declining_stocks: int
    total_volume: int
    market_breadth: float  # Advancing/Declining ratio
    fear_greed_index: str  # FEAR, NEUTRAL, GREED

@dataclass
class CustomerDataFilter:
    """Customer-specific data filtering preferences"""
    customer_id: str
    customer_name: str
    preferred_sectors: List[str]
    risk_tolerance: str  # LOW, MEDIUM, HIGH
    investment_horizon: str  # SHORT, MEDIUM, LONG
    capital_amount: float

class ExcelDataReporter:
    """Handles Excel report generation for pure data analysis"""
    
    def __init__(self, reports_dir: str = "monitor_data"):
        # Create monitor_data directory - navigate from src/monitor/ to project root
        current_file = os.path.abspath(__file__)
        project_root = os.path.dirname(os.path.dirname(os.path.dirname(current_file)))
        self.reports_dir = os.path.join(project_root, "data", reports_dir)
        os.makedirs(self.reports_dir, exist_ok=True)
        
        # Color scheme for data visualization
        self.colors = {
            'header': PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid'),
            'positive': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
            'negative': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
            'neutral': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
            'strong_up': PatternFill(start_color='00B050', end_color='00B050', fill_type='solid'),
            'strong_down': PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid'),
            'sector_tech': PatternFill(start_color='D4E6F1', end_color='D4E6F1', fill_type='solid'),
            'sector_banking': PatternFill(start_color='D5E8D4', end_color='D5E8D4', fill_type='solid'),
            'sector_pharma': PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid'),
            'sector_auto': PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'),
            'sector_energy': PatternFill(start_color='E2E3E5', end_color='E2E3E5', fill_type='solid'),
            'sector_fmcg': PatternFill(start_color='E1D5E7', end_color='E1D5E7', fill_type='solid')
        }
        
        self.fonts = {
            'header': Font(color='FFFFFF', bold=True, size=12),
            'title': Font(bold=True, size=14),
            'subtitle': Font(bold=True, size=11),
            'normal': Font(size=10),
            'data': Font(size=9)
        }
    
    def create_comprehensive_data_report(self, market_overview: MarketOverview, 
                                       stock_data: List[StockData],
                                       technical_indicators: List[TechnicalIndicators],
                                       sector_analysis: List[SectorAnalysis],
                                       customer_filter: CustomerDataFilter) -> tuple[str, str]:
        """Create separate market and customer Excel reports"""
        
        date_str = market_overview.analysis_date.strftime('%Y%m%d')
        
        # Create general market data file
        market_filename = f"{self.reports_dir}/MarketData_General_{date_str}.xlsx"
        market_wb = Workbook()
        market_wb.remove(market_wb.active)
        
        self._create_general_market_overview_sheet(market_wb, market_overview)
        self._create_stock_data_sheet(market_wb, stock_data)
        self._create_technical_indicators_sheet(market_wb, technical_indicators)
        self._create_sector_analysis_sheet(market_wb, sector_analysis)
        self._create_historical_trends_sheet(market_wb, stock_data)
        
        market_wb.save(market_filename)
        logger.info(f"General market report saved: {market_filename}")
        
        # Create customer-specific file in customer folder
        current_file = os.path.abspath(__file__)
        project_root = os.path.dirname(os.path.dirname(os.path.dirname(current_file)))
        customer_dir = os.path.join(project_root, "data", "monitor_data", customer_filter.customer_id)
        os.makedirs(customer_dir, exist_ok=True)
        
        customer_filename = f"{customer_dir}/CustomerData_{customer_filter.customer_id}_{date_str}.xlsx"
        customer_wb = Workbook()
        customer_wb.remove(customer_wb.active)
        
        self._create_customer_overview_sheet(customer_wb, market_overview, customer_filter)
        self._create_filtered_data_sheet(customer_wb, stock_data, technical_indicators, customer_filter)
        
        customer_wb.save(customer_filename)
        logger.info(f"Customer report saved: {customer_filename}")
        
        return market_filename, customer_filename
    
    def _create_general_market_overview_sheet(self, wb: Workbook, market_overview: MarketOverview):
        """Create general market overview sheet"""
        ws = wb.create_sheet("📊 Market Overview", 0)
        
        # Title and date
        ws['A1'] = "GENERAL MARKET DATA ANALYSIS"
        ws['A2'] = f"Analysis Date: {market_overview.analysis_date.strftime('%B %d, %Y %H:%M')}"
        ws['A3'] = "Comprehensive Market Analysis - All Sectors"
        
        for row in range(1, 4):
            ws[f'A{row}'].font = self.fonts['title']
        
        # Market Sentiment Section
        ws['A5'] = "📈 MARKET SENTIMENT"
        ws['A5'].font = self.fonts['subtitle']
        
        sentiment_color = self.colors['positive'] if market_overview.market_sentiment == 'BULLISH' else \
                         self.colors['negative'] if market_overview.market_sentiment == 'BEARISH' else \
                         self.colors['neutral']
        
        sentiment_data = [
            ['Overall Market Sentiment', market_overview.market_sentiment],
            ['Market Volatility', f"{market_overview.market_volatility:.2%}"],
            ['Fear/Greed Index', market_overview.fear_greed_index],
            ['Market Breadth Ratio', f"{market_overview.market_breadth:.2f}"]
        ]
        
        for i, (metric, value) in enumerate(sentiment_data):
            ws[f'A{7+i}'] = metric
            cell = ws[f'B{7+i}']
            cell.value = value
            if i == 0:
                cell.fill = sentiment_color
                cell.font = Font(bold=True)
        
        # Market Statistics
        ws['A13'] = "📊 MARKET STATISTICS"
        ws['A13'].font = self.fonts['subtitle']
        
        stats_data = [
            ['Advancing Stocks', market_overview.advancing_stocks],
            ['Declining Stocks', market_overview.declining_stocks],
            ['Total Trading Volume', f"{market_overview.total_volume:,}"],
            ['Net Advancers', market_overview.advancing_stocks - market_overview.declining_stocks]
        ]
        
        for i, (metric, value) in enumerate(stats_data):
            ws[f'A{15+i}'] = metric
            ws[f'B{15+i}'] = value
        
        self._auto_adjust_columns(ws)
    
    def _create_customer_overview_sheet(self, wb: Workbook, market_overview: MarketOverview, 
                                      customer_filter: CustomerDataFilter):
        """Create customer-specific overview sheet"""
        ws = wb.create_sheet("🎯 Customer Overview", 0)
        
        # Title and date
        ws['A1'] = f"CUSTOMER PORTFOLIO ANALYSIS - {customer_filter.customer_name}"
        ws['A2'] = f"Analysis Date: {market_overview.analysis_date.strftime('%B %d, %Y %H:%M')}"
        ws['A3'] = f"Customer ID: {customer_filter.customer_id}"
        
        for row in range(1, 4):
            ws[f'A{row}'].font = self.fonts['title']
        
        # Customer Profile Section
        ws['A5'] = "👤 CUSTOMER PROFILE"
        ws['A5'].font = self.fonts['subtitle']
        
        profile_data = [
            ['Risk Tolerance', customer_filter.risk_tolerance],
            ['Investment Horizon', customer_filter.investment_horizon],
            ['Capital Amount', f"₹{customer_filter.capital_amount:,.0f}"],
            ['Preferred Sectors', ', '.join(customer_filter.preferred_sectors)]
        ]
        
        for i, (metric, value) in enumerate(profile_data):
            ws[f'A{7+i}'] = metric
            ws[f'B{7+i}'] = value
        
        # Market Context for Customer
        ws['A13'] = "📈 MARKET CONTEXT"
        ws['A13'].font = self.fonts['subtitle']
        
        context_data = [
            ['Market Sentiment', market_overview.market_sentiment],
            ['Market Volatility', f"{market_overview.market_volatility:.2%}"],
            ['Market Breadth', f"{market_overview.market_breadth:.2f}"]
        ]
        
        for i, (metric, value) in enumerate(context_data):
            ws[f'A{15+i}'] = metric
            ws[f'B{15+i}'] = value
        
        self._auto_adjust_columns(ws)
    
    def _create_stock_data_sheet(self, wb: Workbook, stock_data: List[StockData]):
        """Create comprehensive stock data sheet"""
        ws = wb.create_sheet("💰 Stock Data")
        
        # Title
        ws['A1'] = "COMPREHENSIVE STOCK DATA"
        ws['A1'].font = self.fonts['title']
        
        # Create DataFrame from stock data
        data_rows = []
        for stock in stock_data:
            data_rows.append({
                'Symbol': stock.symbol,
                'Company Name': stock.company_name,
                'Sector': stock.sector,
                'Current Price (₹)': stock.current_price,
                'Open (₹)': stock.open_price,
                'High (₹)': stock.high_price,
                'Low (₹)': stock.low_price,
                'Volume': stock.volume,
                'Day Change (₹)': stock.price_change,
                'Day Change %': stock.price_change_pct,
                'Market Cap (₹Cr)': stock.market_cap / 10000000 if stock.market_cap else 0,
                'P/E Ratio': stock.pe_ratio,
                'Date': stock.date
            })
        
        if not data_rows:
            ws['A3'] = "No stock data available"
            return
        
        df = pd.DataFrame(data_rows)
        df = df.sort_values(['Sector', 'Day Change %'], ascending=[True, False])
        
        # Add headers
        headers = list(df.columns)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.fill = self.colors['header']
            cell.font = self.fonts['header']
        
        # Add data with sector-based coloring
        sector_colors = {
            'TECH': self.colors['sector_tech'],
            'BANKING': self.colors['sector_banking'], 
            'PHARMA': self.colors['sector_pharma'],
            'AUTO': self.colors['sector_auto'],
            'ENERGY': self.colors['sector_energy'],
            'FMCG': self.colors['sector_fmcg']
        }
        
        for row_idx, (_, row) in enumerate(df.iterrows(), 4):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Color code by sector
                sector = row['Sector']
                if sector in sector_colors:
                    cell.fill = sector_colors[sector]
                
                # Highlight day change
                if col_idx == df.columns.get_loc('Day Change %') + 1:
                    if value > 0:
                        cell.fill = self.colors['positive']
                    elif value < 0:
                        cell.fill = self.colors['negative']
        
        self._auto_adjust_columns(ws)
    
    def _create_technical_indicators_sheet(self, wb: Workbook, technical_indicators: List[TechnicalIndicators]):
        """Create technical indicators analysis sheet"""
        ws = wb.create_sheet("🔧 Technical Indicators")
        
        ws['A1'] = "TECHNICAL ANALYSIS INDICATORS"
        ws['A1'].font = self.fonts['title']
        
        # Create DataFrame
        indicator_rows = []
        for indicator in technical_indicators:
            indicator_rows.append({
                'Symbol': indicator.symbol,
                'RSI (14)': indicator.rsi,
                'SMA 20': indicator.sma_20,
                'SMA 50': indicator.sma_50,
                'SMA 200': indicator.sma_200,
                'Volatility (30d)': indicator.volatility,
                'Momentum 20d': indicator.momentum_20d,
                'Momentum 50d': indicator.momentum_50d,
                'Bollinger Upper': indicator.bollinger_upper,
                'Bollinger Lower': indicator.bollinger_lower,
                'MACD': indicator.macd,
                'Signal Strength': indicator.signal_strength
            })
        
        if not indicator_rows:
            ws['A3'] = "No technical indicators available"
            return
        
        df = pd.DataFrame(indicator_rows)
        
        # Add headers
        for col, header in enumerate(df.columns, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.fill = self.colors['header']
            cell.font = self.fonts['header']
        
        # Add data with indicator-based coloring
        for row_idx, (_, row) in enumerate(df.iterrows(), 4):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Color code RSI
                if col_idx == df.columns.get_loc('RSI (14)') + 1:
                    if value > 70:
                        cell.fill = self.colors['strong_down']  # Overbought
                    elif value < 30:
                        cell.fill = self.colors['strong_up']    # Oversold
                    else:
                        cell.fill = self.colors['neutral']
                
                # Color code signal strength
                elif col_idx == df.columns.get_loc('Signal Strength') + 1:
                    if 'STRONG_UP' in str(value):
                        cell.fill = self.colors['strong_up']
                    elif 'STRONG_DOWN' in str(value):
                        cell.fill = self.colors['strong_down']
                    elif 'UP' in str(value):
                        cell.fill = self.colors['positive']
                    elif 'DOWN' in str(value):
                        cell.fill = self.colors['negative']
                    else:
                        cell.fill = self.colors['neutral']
        
        self._auto_adjust_columns(ws)
    
    def _create_sector_analysis_sheet(self, wb: Workbook, sector_analysis: List[SectorAnalysis]):
        """Create sector performance analysis sheet"""
        ws = wb.create_sheet("🏭 Sector Analysis")
        
        ws['A1'] = "SECTOR PERFORMANCE ANALYSIS"
        ws['A1'].font = self.fonts['title']
        
        # Create DataFrame
        sector_rows = []
        for sector in sector_analysis:
            sector_rows.append({
                'Sector': sector.sector,
                'Stock Count': sector.stock_count,
                'Avg Price Change %': sector.avg_price_change,
                'Avg Volume Change %': sector.avg_volume_change,
                'Sector Volatility': sector.sector_volatility,
                'Trend Direction': sector.trend_direction,
                'Sector Strength': sector.sector_strength,
                'Top Performers': ', '.join(sector.top_performers),
                'Bottom Performers': ', '.join(sector.bottom_performers)
            })
        
        if not sector_rows:
            ws['A3'] = "No sector analysis available"
            return
        
        df = pd.DataFrame(sector_rows)
        df = df.sort_values('Avg Price Change %', ascending=False)
        
        # Add headers
        for col, header in enumerate(df.columns, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.fill = self.colors['header']
            cell.font = self.fonts['header']
        
        # Add data with performance-based coloring
        for row_idx, (_, row) in enumerate(df.iterrows(), 4):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Color code price change
                if col_idx == df.columns.get_loc('Avg Price Change %') + 1:
                    if value > 2:
                        cell.fill = self.colors['strong_up']
                    elif value > 0:
                        cell.fill = self.colors['positive']
                    elif value < -2:
                        cell.fill = self.colors['strong_down']
                    elif value < 0:
                        cell.fill = self.colors['negative']
                    else:
                        cell.fill = self.colors['neutral']
                
                # Color code trend direction
                elif col_idx == df.columns.get_loc('Trend Direction') + 1:
                    if value == 'UPTREND':
                        cell.fill = self.colors['positive']
                    elif value == 'DOWNTREND':
                        cell.fill = self.colors['negative']
                    else:
                        cell.fill = self.colors['neutral']
        
        self._auto_adjust_columns(ws)
    
    def _create_historical_trends_sheet(self, wb: Workbook, stock_data: List[StockData]):
        """Create historical trends and patterns sheet"""
        ws = wb.create_sheet("📈 Historical Trends")
        
        ws['A1'] = "HISTORICAL TRENDS & PATTERNS"
        ws['A1'].font = self.fonts['title']
        
        # For now, create a summary of current vs historical performance
        ws['A3'] = "📊 Price Movement Summary"
        ws['A3'].font = self.fonts['subtitle']
        
        # Analyze current data for trends
        sectors = {}
        for stock in stock_data:
            if stock.sector not in sectors:
                sectors[stock.sector] = {'stocks': [], 'avg_change': 0}
            sectors[stock.sector]['stocks'].append(stock)
        
        # Calculate sector averages
        summary_data = []
        for sector, data in sectors.items():
            stocks = data['stocks']
            avg_change = np.mean([stock.price_change_pct for stock in stocks])
            volatility = np.std([stock.price_change_pct for stock in stocks])
            
            summary_data.append({
                'Sector': sector,
                'Stock Count': len(stocks),
                'Average Change %': avg_change,
                'Price Volatility': volatility,
                'Best Performer': max(stocks, key=lambda x: x.price_change_pct).symbol,
                'Worst Performer': min(stocks, key=lambda x: x.price_change_pct).symbol
            })
        
        df = pd.DataFrame(summary_data)
        
        # Add to worksheet
        for col, header in enumerate(df.columns, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.fill = self.colors['header']
            cell.font = self.fonts['header']
        
        for row_idx, (_, row) in enumerate(df.iterrows(), 6):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Add trend analysis notes
        ws['A15'] = "📝 TREND ANALYSIS NOTES"
        ws['A15'].font = self.fonts['subtitle']
        
        notes = [
            "• This analysis is based on current day's data",
            "• For comprehensive historical analysis, connect to historical database",
            "• Trends are identified using technical indicators and price movements",
            "• Sector rotation patterns can be observed from relative performance"
        ]
        
        for i, note in enumerate(notes):
            ws[f'A{17+i}'] = note
        
        self._auto_adjust_columns(ws)
    
    def _create_filtered_data_sheet(self, wb: Workbook, stock_data: List[StockData], 
                                  technical_indicators: List[TechnicalIndicators],
                                  customer_filter: CustomerDataFilter):
        """Create customer-specific filtered data sheet"""
        ws = wb.create_sheet("🎯 Customer Filtered Data")
        
        ws['A1'] = f"FILTERED DATA FOR {customer_filter.customer_name}"
        ws['A2'] = f"Sectors: {', '.join(customer_filter.preferred_sectors)} | Risk: {customer_filter.risk_tolerance}"
        
        ws['A1'].font = self.fonts['title']
        ws['A2'].font = self.fonts['subtitle']
        
        # Filter stocks by customer preferences
        filtered_stocks = [stock for stock in stock_data 
                          if stock.sector in customer_filter.preferred_sectors]
        
        if not filtered_stocks:
            ws['A4'] = "No stocks found matching customer preferences"
            return
        
        # Create comprehensive filtered view
        filtered_data = []
        for stock in filtered_stocks:
            # Find corresponding technical indicators
            tech_data = next((t for t in technical_indicators if t.symbol == stock.symbol), None)
            
            row_data = {
                'Symbol': stock.symbol,
                'Company': stock.company_name,
                'Sector': stock.sector,
                'Price (₹)': stock.current_price,
                'Change %': stock.price_change_pct,
                'Volume': stock.volume,
                'Market Cap (₹Cr)': stock.market_cap / 10000000 if stock.market_cap else 0,
                'RSI': tech_data.rsi if tech_data else 'N/A',
                'Volatility': tech_data.volatility if tech_data else 'N/A',
                'Signal': tech_data.signal_strength if tech_data else 'N/A',
                'Risk Level': self._assess_risk_level(stock, tech_data, customer_filter.risk_tolerance)
            }
            filtered_data.append(row_data)
        
        # Sort by performance
        filtered_df = pd.DataFrame(filtered_data)
        filtered_df = filtered_df.sort_values('Change %', ascending=False)
        
        # Add headers
        for col, header in enumerate(filtered_df.columns, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.fill = self.colors['header']
            cell.font = self.fonts['header']
        
        # Add data with risk-based coloring
        for row_idx, (_, row) in enumerate(filtered_df.iterrows(), 5):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Color code by risk level
                if col_idx == filtered_df.columns.get_loc('Risk Level') + 1:
                    if value == 'LOW':
                        cell.fill = self.colors['positive']
                    elif value == 'HIGH':
                        cell.fill = self.colors['negative']
                    else:
                        cell.fill = self.colors['neutral']
        
        # Add summary statistics
        ws[f'A{len(filtered_df) + 7}'] = "📊 FILTERED DATA SUMMARY"
        ws[f'A{len(filtered_df) + 7}'].font = self.fonts['subtitle']
        
        summary_row = len(filtered_df) + 9
        ws[f'A{summary_row}'] = f"Total Filtered Stocks: {len(filtered_stocks)}"
        ws[f'A{summary_row + 1}'] = f"Average Performance: {filtered_df['Change %'].mean():.2f}%"
        ws[f'A{summary_row + 2}'] = f"Best Performer: {filtered_df.iloc[0]['Symbol']} ({filtered_df.iloc[0]['Change %']:.2f}%)"
        ws[f'A{summary_row + 3}'] = f"Sectors Covered: {len(set(stock.sector for stock in filtered_stocks))}"
        
        self._auto_adjust_columns(ws)
    
    def _assess_risk_level(self, stock: StockData, tech_data: TechnicalIndicators, 
                          customer_risk_tolerance: str) -> str:
        """Assess risk level for display purposes only"""
        if not tech_data:
            return "UNKNOWN"
        
        # Simple risk assessment based on volatility
        if tech_data.volatility < 0.2:
            return "LOW"
        elif tech_data.volatility < 0.4:
            return "MEDIUM"
        else:
            return "HIGH"
    
    def _auto_adjust_columns(self, ws):
        """Auto-adjust column widths"""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

class IndianStockDataCollector:
    """Refactored Monitor Agent - Pure Data Collection & Analysis"""
    
    def __init__(self):
        self.sectors = {
            "TECH": [
                ("INFY.NS", "Infosys Limited"),
                ("TCS.NS", "Tata Consultancy Services"),
                ("WIPRO.NS", "Wipro Limited"),
                ("HCLTECH.NS", "HCL Technologies"),
                ("TECHM.NS", "Tech Mahindra"),
                ("LTI.NS", "Larsen & Toubro Infotech")
            ],
            "BANKING": [
                ("HDFCBANK.NS", "HDFC Bank"),
                ("ICICIBANK.NS", "ICICI Bank"),
                ("KOTAKBANK.NS", "Kotak Mahindra Bank"),
                ("AXISBANK.NS", "Axis Bank"),
                ("SBIN.NS", "State Bank of India"),
                ("INDUSINDBK.NS", "IndusInd Bank")
            ],
            "PHARMA": [
                ("SUNPHARMA.NS", "Sun Pharmaceutical"),
                ("DRREDDY.NS", "Dr. Reddy's Labs"),
                ("CIPLA.NS", "Cipla Limited"),
                ("DIVISLAB.NS", "Divi's Laboratories"),
                ("BIOCON.NS", "Biocon Limited"),
                ("LUPIN.NS", "Lupin Limited")
            ],
            "AUTO": [
                ("TATAMOTORS.NS", "Tata Motors"),
                ("MARUTI.NS", "Maruti Suzuki"),
                ("M&M.NS", "Mahindra & Mahindra"),
                ("BAJAJ-AUTO.NS", "Bajaj Auto"),
                ("EICHERMOT.NS", "Eicher Motors"),
                ("HEROMOTOCO.NS", "Hero MotoCorp")
            ],
            "ENERGY": [
                ("RELIANCE.NS", "Reliance Industries"),
                ("ONGC.NS", "Oil & Natural Gas Corp"),
                ("NTPC.NS", "NTPC Limited"),
                ("TATAPOWER.NS", "Tata Power"),
                ("ADANIGREEN.NS", "Adani Green Energy"),
                ("IOC.NS", "Indian Oil Corporation")
            ],
            "FMCG": [
                ("HINDUNILVR.NS", "Hindustan Unilever"),
                ("ITC.NS", "ITC Limited"),
                ("NESTLEIND.NS", "Nestle India"),
                ("BRITANNIA.NS", "Britannia Industries"),
                ("DABUR.NS", "Dabur India"),
                ("GODREJCP.NS", "Godrej Consumer Products")
            ]
        }
        
        # Create stock mappings
        self.all_stocks = []
        self.stock_names = {}
        self.stock_sectors = {}
        
        for sector, stocks in self.sectors.items():
            for symbol, name in stocks:
                self.all_stocks.append(symbol)
                self.stock_names[symbol] = name
                self.stock_sectors[symbol] = sector
        
        self.excel_reporter = ExcelDataReporter()
        
        logger.info(f"Initialized Stock Data Collector with {len(self.all_stocks)} stocks across {len(self.sectors)} sectors")
    
    def collect_stock_data(self, symbols: List[str]) -> List[StockData]:
        """Collect comprehensive stock data"""
        stock_data = []
        
        logger.info(f"Collecting data for {len(symbols)} stocks...")
        
        for symbol in symbols:
            try:
                stock = yf.Ticker(symbol)
                hist_data = stock.history(period="5d")
                
                if not hist_data.empty and len(hist_data) >= 2:
                    latest = hist_data.iloc[-1]
                    previous = hist_data.iloc[-2] if len(hist_data) > 1 else latest
                    
                    # Calculate changes
                    price_change = latest['Close'] - previous['Close']
                    price_change_pct = (price_change / previous['Close']) * 100
                    
                    # Get additional info
                    try:
                        info = stock.info
                        market_cap = info.get('marketCap', 0)
                        pe_ratio = info.get('forwardPE', 0)
                    except:
                        market_cap = 0
                        pe_ratio = 0
                    
                    stock_data_obj = StockData(
                        symbol=symbol,
                        company_name=self.stock_names.get(symbol, symbol.replace('.NS', '')),
                        sector=self.stock_sectors.get(symbol, 'UNKNOWN'),
                        current_price=float(latest['Close']),
                        price_change=float(price_change),
                        price_change_pct=float(price_change_pct),
                        volume=int(latest['Volume']),
                        market_cap=float(market_cap),
                        pe_ratio=float(pe_ratio) if pe_ratio else 0,
                        open_price=float(latest['Open']),
                        high_price=float(latest['High']),
                        low_price=float(latest['Low']),
                        date=latest.name.strftime('%Y-%m-%d')
                    )
                    
                    stock_data.append(stock_data_obj)
                    logger.info(f"✓ Collected data for {symbol}")
                
            except Exception as e:
                logger.error(f"✗ Error collecting data for {symbol}: {e}")
                continue
        
        logger.info(f"Successfully collected data for {len(stock_data)} stocks")
        return stock_data
    
    def calculate_technical_indicators(self, symbols: List[str]) -> List[TechnicalIndicators]:
        """Calculate comprehensive technical indicators"""
        technical_indicators = []
        
        logger.info(f"Calculating technical indicators for {len(symbols)} stocks...")
        
        for symbol in symbols:
            try:
                stock = yf.Ticker(symbol)
                hist_data = stock.history(period="1y")  # Need more data for indicators
                
                if len(hist_data) < 50:  # Need minimum data for calculations
                    continue
                
                prices = hist_data['Close'].values
                volumes = hist_data['Volume'].values
                
                # Calculate indicators
                rsi = self._calculate_rsi(prices)
                sma_20 = np.mean(prices[-20:])
                sma_50 = np.mean(prices[-50:]) if len(prices) >= 50 else sma_20
                sma_200 = np.mean(prices[-200:]) if len(prices) >= 200 else sma_50
                
                volatility = self._calculate_volatility(prices)
                momentum_20d = self._calculate_momentum(prices, 20)
                momentum_50d = self._calculate_momentum(prices, 50)
                
                bollinger_upper, bollinger_lower = self._calculate_bollinger_bands(prices)
                macd = self._calculate_macd(prices)
                signal_strength = self._determine_signal_strength(prices, rsi, macd)
                
                tech_indicator = TechnicalIndicators(
                    symbol=symbol,
                    rsi=rsi,
                    sma_20=sma_20,
                    sma_50=sma_50,
                    sma_200=sma_200,
                    volatility=volatility,
                    momentum_20d=momentum_20d,
                    momentum_50d=momentum_50d,
                    bollinger_upper=bollinger_upper,
                    bollinger_lower=bollinger_lower,
                    macd=macd,
                    signal_strength=signal_strength
                )
                
                technical_indicators.append(tech_indicator)
                logger.info(f"✓ Calculated indicators for {symbol}")
                
            except Exception as e:
                logger.error(f"✗ Error calculating indicators for {symbol}: {e}")
                continue
        
        logger.info(f"Calculated indicators for {len(technical_indicators)} stocks")
        return technical_indicators
    
    def analyze_sectors(self, stock_data: List[StockData]) -> List[SectorAnalysis]:
        """Analyze sector performance"""
        sector_groups = {}
        
        # Group stocks by sector
        for stock in stock_data:
            if stock.sector not in sector_groups:
                sector_groups[stock.sector] = []
            sector_groups[stock.sector].append(stock)
        
        sector_analyses = []
        
        for sector, stocks in sector_groups.items():
            if not stocks:
                continue
            
            # Calculate sector metrics
            price_changes = [stock.price_change_pct for stock in stocks]
            volumes = [stock.volume for stock in stocks]
            
            avg_price_change = np.mean(price_changes)
            avg_volume_change = 0  # Would need historical volume data
            sector_volatility = np.std(price_changes)
            
            # Determine trend
            positive_count = len([p for p in price_changes if p > 0])
            negative_count = len([p for p in price_changes if p < 0])
            
            if positive_count > negative_count * 1.5:
                trend_direction = "UPTREND"
            elif negative_count > positive_count * 1.5:
                trend_direction = "DOWNTREND"
            else:
                trend_direction = "SIDEWAYS"
            
            # Find top and bottom performers
            sorted_stocks = sorted(stocks, key=lambda x: x.price_change_pct, reverse=True)
            top_performers = [stock.symbol for stock in sorted_stocks[:3]]
            bottom_performers = [stock.symbol for stock in sorted_stocks[-3:]]
            
            # Determine sector strength
            if avg_price_change > 2:
                sector_strength = "STRONG"
            elif avg_price_change > 0:
                sector_strength = "MODERATE"
            else:
                sector_strength = "WEAK"
            
            sector_analysis = SectorAnalysis(
                sector=sector,
                stock_count=len(stocks),
                avg_price_change=avg_price_change,
                avg_volume_change=avg_volume_change,
                sector_volatility=sector_volatility,
                trend_direction=trend_direction,
                top_performers=top_performers,
                bottom_performers=bottom_performers,
                sector_strength=sector_strength
            )
            
            sector_analyses.append(sector_analysis)
        
        logger.info(f"Analyzed {len(sector_analyses)} sectors")
        return sector_analyses
    
    def assess_market_overview(self, stock_data: List[StockData]) -> MarketOverview:
        """Assess overall market conditions"""
        if not stock_data:
            return MarketOverview(
                analysis_date=datetime.now(),
                market_sentiment="NEUTRAL",
                market_volatility=0.0,
                advancing_stocks=0,
                declining_stocks=0,
                total_volume=0,
                market_breadth=1.0,
                fear_greed_index="NEUTRAL"
            )
        
        # Calculate market metrics
        advancing_stocks = len([s for s in stock_data if s.price_change_pct > 0])
        declining_stocks = len([s for s in stock_data if s.price_change_pct < 0])
        total_volume = sum(s.volume for s in stock_data)
        
        market_breadth = advancing_stocks / declining_stocks if declining_stocks > 0 else 2.0
        
        # Calculate market volatility
        price_changes = [s.price_change_pct for s in stock_data]
        market_volatility = np.std(price_changes) / 100
        
        # Determine market sentiment
        avg_change = np.mean(price_changes)
        if avg_change > 1.0 and market_breadth > 1.2:
            market_sentiment = "BULLISH"
        elif avg_change < -1.0 and market_breadth < 0.8:
            market_sentiment = "BEARISH"
        else:
            market_sentiment = "NEUTRAL"
        
        # Simple fear/greed assessment
        if market_volatility > 0.03:
            fear_greed_index = "FEAR"
        elif market_volatility < 0.015 and avg_change > 0.5:
            fear_greed_index = "GREED"
        else:
            fear_greed_index = "NEUTRAL"
        
        return MarketOverview(
            analysis_date=datetime.now(),
            market_sentiment=market_sentiment,
            market_volatility=market_volatility,
            advancing_stocks=advancing_stocks,
            declining_stocks=declining_stocks,
            total_volume=total_volume,
            market_breadth=market_breadth,
            fear_greed_index=fear_greed_index
        )
    
    def run_comprehensive_data_collection(self, customer_filter: CustomerDataFilter) -> str:
        """Run complete data collection and analysis"""
        logger.info("Starting comprehensive market data collection...")
        
        try:
            # Step 1: Collect stock data
            logger.info("Step 1: Collecting stock price data...")
            all_stock_data = self.collect_stock_data(self.all_stocks)
            
            # Step 2: Calculate technical indicators
            logger.info("Step 2: Calculating technical indicators...")
            technical_indicators = self.calculate_technical_indicators(self.all_stocks)
            
            # Step 3: Analyze sectors
            logger.info("Step 3: Analyzing sector performance...")
            sector_analyses = self.analyze_sectors(all_stock_data)
            
            # Step 4: Assess market overview
            logger.info("Step 4: Assessing market conditions...")
            market_overview = self.assess_market_overview(all_stock_data)
            
            # Step 5: Generate Excel reports
            logger.info("Step 5: Generating market and customer reports...")
            market_report, customer_report = self.excel_reporter.create_comprehensive_data_report(
                market_overview=market_overview,
                stock_data=all_stock_data,
                technical_indicators=technical_indicators,
                sector_analysis=sector_analyses,
                customer_filter=customer_filter
            )
            
            # Step 6: Create data export for other agents
            self._export_data_for_agents(all_stock_data, technical_indicators, 
                                       sector_analyses, market_overview, customer_filter)
            
            logger.info("Comprehensive data collection completed successfully!")
            return market_report, customer_report
            
        except Exception as e:
            logger.error(f"Error in data collection: {e}")
            raise
    
    def _export_data_for_agents(self, stock_data: List[StockData], 
                               technical_indicators: List[TechnicalIndicators],
                               sector_analyses: List[SectorAnalysis],
                               market_overview: MarketOverview,
                               customer_filter: CustomerDataFilter):
        """Export structured data for other agents to use"""
        
        # Create data export directory in customer folder
        current_file = os.path.abspath(__file__)
        project_root = os.path.dirname(os.path.dirname(os.path.dirname(current_file)))
        export_dir = os.path.join(project_root, "data", "monitor_data", customer_filter.customer_id)
        os.makedirs(export_dir, exist_ok=True)
        
        # Export for Investment Planner
        planner_data = {
            "timestamp": datetime.now().isoformat(),
            "customer_id": customer_filter.customer_id,
            "market_overview": {
                "sentiment": market_overview.market_sentiment,
                "volatility": market_overview.market_volatility,
                "breadth": market_overview.market_breadth
            },
            "stocks": [],
            "sectors": [],
            "customer_preferences": {
                "sectors": customer_filter.preferred_sectors,
                "risk_tolerance": customer_filter.risk_tolerance,
                "investment_horizon": customer_filter.investment_horizon
            }
        }
        
        # Add stock data
        for stock in stock_data:
            if stock.sector in customer_filter.preferred_sectors:
                tech_data = next((t for t in technical_indicators if t.symbol == stock.symbol), None)
                
                stock_entry = {
                    "symbol": stock.symbol,
                    "company_name": stock.company_name,
                    "sector": stock.sector,
                    "current_price": stock.current_price,
                    "price_change_pct": stock.price_change_pct,
                    "volume": stock.volume,
                    "market_cap": stock.market_cap,
                    "technical_indicators": {
                        "rsi": tech_data.rsi if tech_data else None,
                        "volatility": tech_data.volatility if tech_data else None,
                        "momentum_20d": tech_data.momentum_20d if tech_data else None,
                        "signal_strength": tech_data.signal_strength if tech_data else None
                    }
                }
                planner_data["stocks"].append(stock_entry)
        
        # Add sector data
        for sector in sector_analyses:
            if sector.sector in customer_filter.preferred_sectors:
                sector_entry = {
                    "sector": sector.sector,
                    "performance": sector.avg_price_change,
                    "volatility": sector.sector_volatility,
                    "trend": sector.trend_direction,
                    "strength": sector.sector_strength
                }
                planner_data["sectors"].append(sector_entry)
        
        # Save for Investment Planner
        with open(f"{export_dir}/investment_planner_data.json", 'w') as f:
            json.dump(planner_data, f, indent=2)
        
        # Export CSV for backup
        stock_df = pd.DataFrame([{
            'Symbol': s.symbol,
            'Company': s.company_name,
            'Sector': s.sector,
            'Price': s.current_price,
            'Change_Pct': s.price_change_pct,
            'Volume': s.volume,
            'Market_Cap': s.market_cap
        } for s in stock_data])
        
        stock_df.to_csv(f"{export_dir}/stock_data.csv", index=False)
        
        logger.info(f"Data exported for other agents in {export_dir}")
    
    # Technical calculation helper methods
    def _calculate_rsi(self, prices: np.ndarray, period: int = 14) -> float:
        """Calculate RSI"""
        if len(prices) < period + 1:
            return 50.0
        
        deltas = np.diff(prices)
        gains = np.where(deltas > 0, deltas, 0)
        losses = np.where(deltas < 0, -deltas, 0)
        
        avg_gain = np.mean(gains[-period:])
        avg_loss = np.mean(losses[-period:])
        
        if avg_loss == 0:
            return 100.0
        
        rs = avg_gain / avg_loss
        rsi = 100 - (100 / (1 + rs))
        return float(rsi)
    
    def _calculate_volatility(self, prices: np.ndarray, window: int = 30) -> float:
        """Calculate annualized volatility"""
        if len(prices) < 2:
            return 0.0
        
        returns = np.diff(np.log(prices))
        if len(returns) < window:
            volatility = np.std(returns) * np.sqrt(252)
        else:
            volatility = np.std(returns[-window:]) * np.sqrt(252)
        
        return float(volatility)
    
    def _calculate_momentum(self, prices: np.ndarray, period: int) -> float:
        """Calculate price momentum"""
        if len(prices) < period + 1:
            return 0.0
        
        momentum = (prices[-1] / prices[-period-1]) - 1
        return float(momentum)
    
    def _calculate_bollinger_bands(self, prices: np.ndarray, period: int = 20, std_dev: float = 2.0) -> Tuple[float, float]:
        """Calculate Bollinger Bands"""
        if len(prices) < period:
            return float(prices[-1]), float(prices[-1])
        
        sma = np.mean(prices[-period:])
        std = np.std(prices[-period:])
        
        upper_band = sma + (std_dev * std)
        lower_band = sma - (std_dev * std)
        
        return float(upper_band), float(lower_band)
    
    def _calculate_macd(self, prices: np.ndarray, fast: int = 12, slow: int = 26) -> float:
        """Calculate MACD"""
        if len(prices) < slow:
            return 0.0
        
        ema_fast = self._calculate_ema(prices, fast)
        ema_slow = self._calculate_ema(prices, slow)
        
        macd = ema_fast - ema_slow
        return float(macd)
    
    def _calculate_ema(self, prices: np.ndarray, period: int) -> float:
        """Calculate Exponential Moving Average"""
        if len(prices) < period:
            return float(np.mean(prices))
        
        multiplier = 2.0 / (period + 1)
        ema = prices[0]
        
        for price in prices[1:]:
            ema = (price * multiplier) + (ema * (1 - multiplier))
        
        return float(ema)
    
    def _determine_signal_strength(self, prices: np.ndarray, rsi: float, macd: float) -> str:
        """Determine overall signal strength"""
        signals = []
        
        # RSI signals
        if rsi > 80:
            signals.append(-2)  # Strong overbought
        elif rsi > 70:
            signals.append(-1)  # Overbought
        elif rsi < 20:
            signals.append(2)   # Strong oversold
        elif rsi < 30:
            signals.append(1)   # Oversold
        else:
            signals.append(0)   # Neutral
        
        # MACD signals
        if macd > 0:
            signals.append(1)
        elif macd < 0:
            signals.append(-1)
        else:
            signals.append(0)
        
        # Price trend (simple)
        if len(prices) >= 5:
            recent_trend = (prices[-1] / prices[-5]) - 1
            if recent_trend > 0.02:
                signals.append(1)
            elif recent_trend < -0.02:
                signals.append(-1)
            else:
                signals.append(0)
        
        # Calculate overall signal
        total_signal = sum(signals)
        
        if total_signal >= 3:
            return "STRONG_UP"
        elif total_signal >= 1:
            return "UP"
        elif total_signal <= -3:
            return "STRONG_DOWN"
        elif total_signal <= -1:
            return "DOWN"
        else:
            return "NEUTRAL"

# Usage functions
def create_sample_customer_filter() -> CustomerDataFilter:
    """Create a sample customer filter for testing"""
    return CustomerDataFilter(
        customer_id="CUST_001",
        customer_name="Rajesh Kumar",
        preferred_sectors=["TECH", "BANKING", "PHARMA", "AUTO", "ENERGY", "FMCG"],
        risk_tolerance="MEDIUM",
        investment_horizon="MEDIUM",
        capital_amount=100000
    )

def main():
    """Main function to demonstrate refactored Monitor Agent"""
    logger.info("Starting Refactored Monitor Agent - Data Collection Only")
    
    # Create data collector
    data_collector = IndianStockDataCollector()
    
    # Create sample customer filter
    customer_filter = create_sample_customer_filter()
    
    try:
        # Run comprehensive data collection
        market_report, customer_report = data_collector.run_comprehensive_data_collection(customer_filter)
        
        print("\n" + "="*80)
        print("🎉 MARKET DATA COLLECTION COMPLETE!")
        print("="*80)
        print(f"📊 General Market Report: {market_report}")
        print(f"🎯 Customer Report: {customer_report}")
        print(f"👤 Customer: {customer_filter.customer_name}")
        print(f"🏭 Sectors: {', '.join(customer_filter.preferred_sectors)}")
        print(f"⚠️  Risk Profile: {customer_filter.risk_tolerance}")
        print(f"📅 Investment Horizon: {customer_filter.investment_horizon}")
        
        print("\n📋 Excel Report Contains:")
        print("   📊 Market Overview - Market sentiment & conditions")
        print("   💰 Stock Data - Comprehensive price & volume data")
        print("   🔧 Technical Indicators - RSI, moving averages, signals")
        print("   🏭 Sector Analysis - Sector performance & trends")
        print("   📈 Historical Trends - Price patterns & analysis")
        print("   🎯 Filtered Data - Customer-specific data view")
        
        print("\n📤 Data Export for Other Agents:")
        print(f"   📁 data/monitor_data/{customer_filter.customer_id}/")
        print("   📄 investment_planner_data.json - Structured data for Investment Planner")
        print("   📄 stock_data.csv - Backup CSV data")
        
        print(f"\n✅ Open reports to view analysis:")
        print(f"   📊 Market: {market_report}")
        print(f"   🎯 Customer: {customer_report}")
        print("="*80)
        
    except Exception as e:
        logger.error(f"Error in main execution: {e}")
        print(f"❌ Error: {e}")

if __name__ == "__main__":
    main()